#!/usr/bin/env python3
"""Generate Job_Search_Tracker.xlsx from the canonical job_search_tracker.csv.

**Regenerate-only, one direction.** The CSV and the per-application `outcome.md`
files are the only writers of truth; this workbook is a read-only view. Nothing
is ever read back out of it, which is the property that makes regeneration safe:
it cannot lose a note, because it never held the only copy of one. Corrections
travel through conversation ("rejected by Acme") into `/outcome`.

    python harness/tracker_xlsx.py
    python harness/tracker_xlsx.py --csv job_search_tracker.csv --out Job_Search_Tracker.xlsx

Four tabs: Applications, Summary, Shortlist, Search Runs.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import os
import sys
from pathlib import Path

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_CSV = os.path.join(ROOT, "job_search_tracker.csv")
DEFAULT_OUT = os.path.join(ROOT, "Job_Search_Tracker.xlsx")
RUN_LOG = os.path.join(ROOT, "run_log.csv")
SHORTLIST = os.path.join(ROOT, "shortlist.csv")

# Folder-name matching is shared with the archiver so the workbook's links and
# the applied/ move can never disagree about which folder a row belongs to.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from archive_applications import RESERVED, match_folder  # noqa: E402
import status  # noqa: E402
import today as today_mod  # noqa: E402
import tracker_row  # noqa: E402

APPLICATIONS_DIR = os.path.join(ROOT, "documents", "applications")

# The canonical tracker header lives in harness/tracker_row.py, which is the
# only writer. It is not duplicated here: the copy that used to sit in this file
# claimed in its comment to include location/rationale/submitted_date and did
# not, and nothing read it.

# Status classification lives in harness/status.py — one vocabulary shared with
# the archiver, so a row written by upstream's /outcome ("applied", "interview")
# lands in the same bucket as one written by /apply-any ("in_progress").
# Previously anything outside these six sets counted as neither Open nor Closed
# and silently disappeared from the funnel.
OPEN_STATUSES = status.OPEN
CLOSED_STATUSES = status.CLOSED
RESPONDED_STATUSES = status.RESPONDED
INTERVIEW_STATUSES = status.REACHED_INTERVIEW

FILLS = {
    "in_progress":    "FFF2CC",  # amber      - live, waiting
    "interview_only": "DDEBF7",  # blue       - talking
    "hired":          "C6EFCE",  # green      - won
    "offer_declined": "E2EFDA",  # pale green - won then declined
    "rejected":       "F8CBAD",  # red        - closed
    "no_response":    "EDEDED",  # grey       - gone quiet
}
FOLLOWUP_DAYS = status.FOLLOWUP_DAYS

# Shortlist verdicts: why a scored candidate did or did not become an application.
SHORTLIST_FILLS = {
    "qualified":    "C6EFCE",  # green - gates passed, package drafted
    "not-drafted":  "FFF2CC",  # amber - gates passed but a stated requirement is unmet
    "not-resolved": "DDEBF7",  # blue  - title-level triage only, still to resolve
    "gate-fail":    "F8CBAD",  # red   - failed a hard constraint, do not revisit
}


def read_csv(path):
    """Read the shortlist. The TRACKER goes through `tracker_row.read_rows`,
    which heals a row carrying an unquoted comma; this reader used to raise
    `AttributeError: 'list' object has no attribute 'strip'` on exactly the
    row the writer accepts.
    """
    if not os.path.exists(path):
        return []
    with open(path, newline="", encoding="utf-8-sig") as handle:
        rows = []
        for row in csv.DictReader(handle):
            row.pop(None, None)
            if any((value or "").strip() for value in row.values()):
                rows.append(row)
        return rows


def parse_date(value):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime((value or "").strip(), fmt).date()
        except ValueError:
            continue
    return None


def _key(row):
    """(company, role) identity shared by the tracker CSV and the shortlist CSV."""
    return ((row.get("company") or "").strip().lower(),
            (row.get("role") or "").strip().lower())


def application_folder(row, root=None):
    """Best-effort path to the application folder, if one exists.

    Submitted applications are moved into documents/applications/applied/ by
    harness/archive_applications.py, so both locations are searched.

    `root` defaults to the repo, but `build()` passes the tracker's own
    directory: resolving against a module global meant a workbook built from a
    CSV anywhere else still linked into the live repo tree, and made the
    workbook tests depend on whatever was in the developer's private
    documents/applications/.
    """
    root = root or ROOT
    applications_dir = os.path.join(root, "documents", "applications")
    if not os.path.isdir(applications_dir):
        return ""
    places = [("documents", "applications")]
    if os.path.isdir(os.path.join(applications_dir, "applied")):
        places.append(("documents", "applications", "applied"))
    for parts in places:
        base = os.path.join(root, *parts)
        names = [n for n in sorted(os.listdir(base))
                 if n not in RESERVED
                 and os.path.isdir(os.path.join(base, n))]
        hit = match_folder(row.get("company", ""), row.get("role", ""), names)
        if hit:
            return os.path.join(*parts, hit)
    return ""


def build(csv_path=DEFAULT_CSV, out_path=DEFAULT_OUT, today=None,
          shortlist_path=None, runlog_path=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        sys.exit("tracker_xlsx: openpyxl is required. "
                 "Run: pip install -r requirements.txt")

    today = today or datetime.date.today()
    # Application folders live beside the tracker, not beside this module.
    tree_root = os.path.dirname(os.path.abspath(csv_path)) or ROOT
    rows, _ = tracker_row.read_rows(Path(csv_path))
    shortlist = shortlist_rows = read_csv(shortlist_path or SHORTLIST)
    rows.sort(key=lambda r: (parse_date(r.get("date")) or datetime.date.min),
              reverse=True)

    workbook = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="44546A")

    def write_sheet(sheet, headers, data_rows):
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font, cell.fill = header_font, header_fill
            cell.alignment = Alignment(vertical="center")
        for row in data_rows:
            sheet.append(row)
        sheet.freeze_panes = "A2"
        if data_rows:
            sheet.auto_filter.ref = "A1:%s%d" % (
                get_column_letter(len(headers)), len(data_rows) + 1
            )
        for index, head in enumerate(headers, 1):
            width = max(
                [len(str(head))]
                + [len(str(r[index - 1])) for r in data_rows if r[index - 1] is not None]
            ) + 2
            sheet.column_dimensions[get_column_letter(index)].width = min(max(width, 10), 52)

    # ------------------------------------------------------------ Applications
    sheet = workbook.active
    sheet.title = "Applications"
    headers = ["Date", "Company", "Role", "Location", "Role type", "Sector", "Channel",
               "Status", "Open/Closed", "Days since", "Follow-up due", "Fit",
               "Rationale", "Contact", "Notes", "CV file", "Cover letter", "Source",
               "Application folder"]
    body, follow_ups = [], 0
    for row in rows:
        row_status = status.normalize(row.get("status"))
        applied_on = parse_date(row.get("date"))
        days = (today - applied_on).days if applied_on else ""
        is_open = status.is_open(row.get("status"))
        # Ask today.py, so the workbook and the daily brief cannot disagree.
        # This column used to count from the application date alone, which
        # meant it kept saying a follow-up was due after one had been sent and
        # logged - the exact complaint /today was fixed for.
        quiet = today_mod.days_quiet(row, today)
        due = "YES" if (is_open and quiet is not None and quiet >= FOLLOWUP_DAYS
                        and today_mod.followups_sent(row) < 2) else ""
        if due:
            follow_ups += 1
        body.append([row.get("date", ""), row.get("company", ""), row.get("role", ""),
                     row.get("location", ""), row.get("role_type", ""),
                     row.get("sector", ""), row.get("channel", ""), row_status,
                     "Open" if is_open else "Closed", days, due,
                     row.get("fit_rating", ""), row.get("rationale", ""),
                     row.get("contact_person", ""), row.get("notes", ""),
                     row.get("cv_file", ""), row.get("cover_letter_file", ""),
                     row.get("source", ""), application_folder(row, tree_root)])
    write_sheet(sheet, headers, body)

    # Derived from the header, never hardcoded: inserting a column used to
    # silently hyperlink the wrong cell, and no test would have caught it.
    source_col = headers.index("Source") + 1
    folder_col = headers.index("Application folder") + 1
    for index, row in enumerate(rows, start=2):
        fill = FILLS.get(status.normalize(row.get("status")))
        if fill:
            pattern = PatternFill("solid", fgColor=fill)
            for column in range(1, len(headers) + 1):
                sheet.cell(row=index, column=column).fill = pattern
        source = (row.get("source") or "").strip()
        if source.startswith("http"):
            cell = sheet.cell(row=index, column=source_col)
            cell.hyperlink, cell.style = source, "Hyperlink"
        folder = sheet.cell(row=index, column=folder_col).value
        if folder:
            cell = sheet.cell(row=index, column=folder_col)
            cell.hyperlink, cell.style = folder.replace("/", os.sep), "Hyperlink"

    # ---------------------------------------------------------------- Summary
    total = len(rows)
    by_status: dict[str, int] = {}
    by_source: dict[str, int] = {}
    for row in rows:
        row_status = status.normalize(row.get("status"))
        by_status[row_status] = by_status.get(row_status, 0) + 1
        key = (row.get("channel") or row.get("source") or "unknown").strip() or "unknown"
        by_source[key] = by_source.get(key, 0) + 1
    responded = sum(by_status.get(s, 0) for s in RESPONDED_STATUSES)
    interviews = sum(by_status.get(s, 0) for s in INTERVIEW_STATUSES)
    dates = [d for d in (parse_date(r.get("date")) for r in rows) if d]
    weeks = max(((max(dates) - min(dates)).days / 7.0), 1.0) if dates else 1.0

    def pct(numerator, denominator):
        return ("%.0f%%" % (100.0 * numerator / denominator)) if denominator else "n/a"

    summary = [
        ["Applications sent", total],
        ["Open (in progress or interviewing)",
         sum(by_status.get(s, 0) for s in OPEN_STATUSES)],
        ["Closed", sum(by_status.get(s, 0) for s in CLOSED_STATUSES)],
        ["", ""],
        ["FUNNEL", ""],
        ["Applied", total],
        ["Got a response", responded],
        ["Reached interview", interviews],
        ["Offers (hired or declined)",
         by_status.get("hired", 0) + by_status.get("offer_declined", 0)],
        ["", ""],
        ["RATES", ""],
        ["Response rate", pct(responded, total)],
        ["Interview conversion (of responses)", pct(interviews, responded)],
        ["Interview rate (of all applications)", pct(interviews, total)],
        ["Applications per week", round(total / weeks, 1) if total else 0],
        ["", ""],
        ["ACTION", ""],
        ["Follow-ups due (open, silent >%d days)" % FOLLOWUP_DAYS, follow_ups],
        ["", ""],
        ["BY STATUS", ""],
    ]
    summary += [[s or "(blank)", n]
                for s, n in sorted(by_status.items(), key=lambda kv: -kv[1])]
    summary += [["", ""], ["BY SOURCE / CHANNEL", ""]]
    summary += [[s, n] for s, n in sorted(by_source.items(), key=lambda kv: -kv[1])]

    # Which channel actually converts. Counts alone cannot answer the most
    # valuable question in a job search - a board can supply half the
    # applications and none of the replies, and raw totals hide that entirely.
    replies: dict[str, int] = {}
    for row in rows:
        key = (row.get("channel") or row.get("source") or "unknown").strip() or "unknown"
        if status.responded(row.get("status")):
            replies[key] = replies.get(key, 0) + 1
    summary += [["", ""], ["RESPONSE RATE BY CHANNEL", "(replies / sent)"]]
    for source, sent in sorted(by_source.items(), key=lambda kv: -kv[1]):
        got = replies.get(source, 0)
        summary.append([source, f"{pct(got, sent)}  ({got}/{sent})"])

    if shortlist_rows:
        verdicts: dict[str, int] = {}
        for row in shortlist_rows:
            verdicts[row.get("verdict", "") or "(blank)"] = verdicts.get(
                row.get("verdict", "") or "(blank)", 0) + 1
        summary += [["", ""], ["SHORTLIST VERDICTS", "(why jobs did not become applications)"]]
        summary += [[v, n] for v, n in sorted(verdicts.items(), key=lambda kv: -kv[1])]
    summary += [["", ""], ["Generated", today.isoformat()],
                ["Source of truth",
                 os.path.basename(csv_path) + " - never edit this workbook by hand"]]
    write_sheet(workbook.create_sheet("Summary"), ["Metric", "Value"], summary)

    # -------------------------------------------------------------- Shortlist
    # Scored candidates that did NOT become applications, so the reasoning is
    # readable here instead of buried in a session transcript. Highest score
    # first; gate failures last.
    shortlist = read_csv(shortlist_path or SHORTLIST)
    def _score(row):
        """Sort key that tolerates "8.5", "n/a" or a stray space.

        `int(...)` here used to raise ValueError and take the whole workbook
        down, while the rendering three lines later carefully checked
        .isdigit() first.
        """
        try:
            return -float(str(row.get("score") or 0).strip())
        except ValueError:
            return 0.0

    shortlist.sort(key=lambda r: (r.get("verdict") == "gate-fail", _score(r)))
    # "In pipeline?" asks whether this candidate became an application, so it is
    # answered from the tracker itself. The folder-name matcher is too fuzzy
    # here: a legal name like "Acme, Baker & Clark LLP" never matches the folder
    # "Acme_...", so the row would silently read "no".
    applied = {_key(r) for r in rows}
    sheet = workbook.create_sheet("Shortlist")
    write_sheet(sheet, ["Date", "Company", "Role", "Location", "Score", "Verdict",
                        "In pipeline?", "Closes", "Source", "Rationale", "Posting"],
                [[r.get("date", ""), r.get("company", ""), r.get("role", ""),
                  r.get("location", ""),
                  int(r["score"]) if (r.get("score") or "").strip().isdigit()
                  else r.get("score", ""),
                  r.get("verdict", ""),
                  "yes" if _key(r) in applied else "no",
                  r.get("deadline", ""),
                  r.get("source", ""), r.get("rationale", ""), r.get("url", "")]
                 for r in shortlist])
    for index, row in enumerate(shortlist, start=2):
        fill = SHORTLIST_FILLS.get(row.get("verdict", ""))
        if fill:
            for cell in sheet[index]:
                cell.fill = PatternFill("solid", fgColor=fill)
        if row.get("url"):
            cell = sheet.cell(row=index, column=11)
            cell.hyperlink, cell.style = row["url"], "Hyperlink"

    # ------------------------------------------------------------ Search Runs
    runs = read_csv(runlog_path or RUN_LOG)
    write_sheet(workbook.create_sheet("Search Runs"),
                ["Date", "Portal", "Query", "Found", "New", "Notes"],
                [[r.get("date", ""), r.get("portal", ""), r.get("query", ""),
                  r.get("found", ""), r.get("new", ""), r.get("notes", "")]
                 for r in runs])

    workbook.save(out_path)
    return out_path, total, follow_ups


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Regenerate the XLSX view of the tracker CSV"
    )
    parser.add_argument("--csv", default=DEFAULT_CSV)
    parser.add_argument("--out", default=DEFAULT_OUT)
    parser.add_argument("--shortlist", default=SHORTLIST)
    parser.add_argument("--runlog", default=RUN_LOG)
    args = parser.parse_args()
    if not os.path.exists(args.csv):
        print("tracker_xlsx: %s does not exist yet - nothing to view." % args.csv)
        print("It is created by /apply-any or /outcome when you record your first "
              "application.")
        sys.exit(0)
    path, total, follow_ups = build(args.csv, args.out,
                                    shortlist_path=args.shortlist,
                                    runlog_path=args.runlog)
    print("Wrote %s - %d application(s), %d follow-up(s) due." % (path, total, follow_ups))
