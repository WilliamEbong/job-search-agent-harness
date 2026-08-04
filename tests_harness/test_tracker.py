"""Tests for the tracking layer: workbook generation, archiver, folder matcher.

The workbook is a *view*. The property that makes it safe to regenerate at will
is that nothing is ever read back out of it, so it can never hold the only copy
of anything. `test_regenerating_cannot_lose_a_csv_note` is the test that would
fail if someone ever made it writable — which is the single change most likely
to be proposed and most likely to destroy data.

The folder matcher is shared between the archiver and the workbook so the two
can never disagree about which folder a tracker row belongs to. Its cases come
from real name-divergence patterns: shortened folder names, reordered words,
ampersands spelled out, and legal-entity suffixes.
"""

from __future__ import annotations

import csv
import datetime
import os
import shutil
import sys
import tempfile
import time
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "harness"))

import archive_applications as archiver  # noqa: E402
import tracker_xlsx  # noqa: E402

HEADER = ["date", "company", "sector", "role", "role_type", "channel", "status",
          "contact_person", "fit_rating", "notes", "cv_file", "cover_letter_file",
          "source", "location", "rationale", "submitted_date"]


def write_tracker(path: Path, rows: list[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=HEADER)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in HEADER})


def demo_rows() -> list[dict]:
    return [
        {"date": "2026-07-01", "company": "Rivermouth Environmental Consulting",
         "role": "Environmental Data Analyst", "status": "in_progress",
         "fit_rating": "82", "notes": "drafted, not yet submitted",
         "source": "https://careers.example.com/rivermouth/4471",
         "location": "Winnipeg, MB", "channel": "company site"},
        {"date": "2026-06-02", "company": "Prairie Grid Utilities",
         "role": "Environmental Compliance Officer", "status": "rejected",
         "fit_rating": "71", "notes": "no interview", "channel": "board",
         "location": "Winnipeg, MB"},
        {"date": "2026-05-20", "company": "Lakeshore Municipal Water Authority",
         "role": "Water Quality Technician", "status": "interview_only",
         "fit_rating": "76", "notes": "second round booked", "channel": "board",
         "location": "Lakeshore, MB"},
    ]


class FolderMatcher(unittest.TestCase):
    def test_matches_a_shortened_folder_name(self):
        folders = ["Rivermouth_Environmental_Data_Analyst"]
        hit = archiver.match_folder(
            "Rivermouth Environmental Consulting",
            "Environmental Data Analyst, Water Programmes", folders,
        )
        self.assertEqual("Rivermouth_Environmental_Data_Analyst", hit)

    def test_matches_despite_reordered_words(self):
        """Word order differs and one word is abbreviated, and it still matches.

        Folder words are {senior, lab, technician}; the row's role supplies
        {technician, senior, laboratory}. "lab" does not equal "laboratory", so
        the score is 2/3 = 0.67 — over the 0.6 threshold. That headroom is the
        point: folder names are abbreviations of posting titles, so an exact
        matcher would miss almost every real pair.
        """
        folders = ["Northwind_Senior_Lab_Technician"]
        hit = archiver.match_folder("Northwind Analytical Services",
                                    "Technician, Senior Laboratory", folders)
        self.assertEqual("Northwind_Senior_Lab_Technician", hit)

    def test_threshold_still_rejects_a_half_match(self):
        """The other side of that headroom: 1 of 3 words is not a match."""
        folders = ["Northwind_Senior_Lab_Technician"]
        hit = archiver.match_folder("Northwind Analytical Services",
                                    "Technician, Field Sampling", folders)
        self.assertEqual("", hit)

    def test_ampersand_spelled_out_in_folder_matches_the_row(self):
        folders = ["Acme_Learning_and_Development_Lead"]
        hit = archiver.match_folder("Acme", "Learning & Development Lead", folders)
        self.assertEqual("Acme_Learning_and_Development_Lead", hit)

    def test_legal_suffix_in_the_company_name_still_matches(self):
        folders = ["Acme_Environmental_Analyst"]
        hit = archiver.match_folder("Acme, Baker & Clark LLP",
                                    "Environmental Analyst", folders)
        self.assertEqual("Acme_Environmental_Analyst", hit)

    def test_ties_go_to_the_more_specific_folder(self):
        folders = ["Acme_Analyst", "Acme_Analyst_Water_Programmes"]
        hit = archiver.match_folder("Acme", "Analyst, Water Programmes", folders)
        self.assertEqual("Acme_Analyst_Water_Programmes", hit)

    def test_wrong_company_never_matches(self):
        folders = ["Northwind_Environmental_Data_Analyst"]
        hit = archiver.match_folder("Rivermouth", "Environmental Data Analyst",
                                    folders)
        self.assertEqual("", hit)

    def test_unrelated_role_at_the_right_company_does_not_match(self):
        folders = ["Acme_Warehouse_Supervisor"]
        hit = archiver.match_folder("Acme", "Environmental Data Analyst", folders)
        self.assertEqual("", hit)

    def test_companies_sharing_a_first_word_do_not_share_a_folder(self):
        """The first word is a filter, not an identity.

        Both companies key on "canadian". With the right folder missing, the
        role score alone put this row in the wrong company's folder:
        {tire, data, analyst} against a role of {data, analyst} is 2/3 = 0.67,
        over the threshold. Landing a row in another employer's folder is worse
        than not matching, because the workbook then links to it and the
        applied/ move carries it there.
        """
        folders = ["Canadian_Tire_Data_Analyst"]
        hit = archiver.match_folder("Canadian Nuclear Laboratories",
                                    "Data Analyst", folders)
        self.assertEqual("", hit)

    def test_the_right_folder_still_wins_when_both_are_present(self):
        folders = ["Canadian_Tire_Data_Analyst",
                   "Canadian_Nuclear_Laboratories_Data_Analyst"]
        hit = archiver.match_folder("Canadian Nuclear Laboratories",
                                    "Data Analyst", folders)
        self.assertEqual("Canadian_Nuclear_Laboratories_Data_Analyst", hit)


class WorkbookGeneration(unittest.TestCase):
    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="harness-tracker-"))
        self.csv_path = self.tmp / "job_search_tracker.csv"
        self.out = self.tmp / "Job_Search_Tracker.xlsx"
        write_tracker(self.csv_path, demo_rows())

    def tearDown(self):
        shutil.rmtree(self.tmp, ignore_errors=True)

    def build(self):
        return tracker_xlsx.build(str(self.csv_path), str(self.out),
                                  today=datetime.date(2026, 8, 4))

    def test_produces_the_four_tabs(self):
        from openpyxl import load_workbook

        self.build()
        workbook = load_workbook(self.out)
        self.assertEqual(["Applications", "Summary", "Shortlist", "Search Runs"],
                         workbook.sheetnames)

    def test_row_count_and_followups(self):
        _, total, follow_ups = self.build()
        self.assertEqual(3, total)
        # Only in_progress/interview_only rows older than 10 days count.
        self.assertEqual(2, follow_ups)

    def test_open_and_closed_are_split(self):
        from openpyxl import load_workbook

        self.build()
        sheet = load_workbook(self.out)["Applications"]
        headers = [c.value for c in sheet[1]]
        column = headers.index("Open/Closed") + 1
        values = {sheet.cell(row=r, column=column).value
                  for r in range(2, sheet.max_row + 1)}
        self.assertEqual({"Open", "Closed"}, values)

    def test_source_url_becomes_a_hyperlink(self):
        from openpyxl import load_workbook

        self.build()
        sheet = load_workbook(self.out)["Applications"]
        headers = [c.value for c in sheet[1]]
        column = headers.index("Source") + 1
        links = [sheet.cell(row=r, column=column).hyperlink
                 for r in range(2, sheet.max_row + 1)]
        self.assertTrue(any(link is not None for link in links))

    def test_regenerating_twice_is_idempotent(self):
        first = self.build()[1:]
        second = self.build()[1:]
        self.assertEqual(first, second)

    def test_regenerating_cannot_lose_a_csv_note(self):
        """The workbook is a view. Regeneration must not touch the CSV at all.

        If the workbook ever became a writer, this is the test that would fail —
        and it is the change most likely to be proposed, because 'let me just
        edit the status in Excel' is the obvious thing to want.
        """
        before = self.csv_path.read_text(encoding="utf-8")
        self.build()
        self.build()
        self.assertEqual(before, self.csv_path.read_text(encoding="utf-8"))
        self.assertIn("second round booked", before)

    def test_missing_csv_exits_cleanly_rather_than_crashing(self):
        import subprocess

        proc = subprocess.run(
            [sys.executable, str(ROOT / "harness" / "tracker_xlsx.py"),
             "--csv", str(self.tmp / "nope.csv"), "--out", str(self.out)],
            capture_output=True, text=True,
        )
        self.assertEqual(0, proc.returncode)
        self.assertIn("nothing to view", proc.stdout)


class Archiver(unittest.TestCase):
    """Exercises the real functions against a temporary applications tree."""

    def setUp(self):
        self.tmp = Path(tempfile.mkdtemp(prefix="harness-archive-"))
        self.applications = self.tmp / "documents" / "applications"
        self.applications.mkdir(parents=True)
        self._saved = (archiver.APPLICATIONS_DIR, archiver.ARCHIVE_DIR,
                       archiver.APPLIED_DIR, archiver.TRACKER_CSV)
        archiver.APPLICATIONS_DIR = self.applications
        archiver.ARCHIVE_DIR = self.applications / "archive"
        archiver.APPLIED_DIR = self.applications / "applied"
        archiver.TRACKER_CSV = self.tmp / "job_search_tracker.csv"

    def tearDown(self):
        (archiver.APPLICATIONS_DIR, archiver.ARCHIVE_DIR,
         archiver.APPLIED_DIR, archiver.TRACKER_CSV) = self._saved
        shutil.rmtree(self.tmp, ignore_errors=True)

    def make_application(self, name: str) -> Path:
        folder = self.applications / name
        folder.mkdir()
        (folder / "job_posting.md").write_text("posting", encoding="utf-8")
        return folder

    def test_recent_folder_is_not_archived(self):
        self.make_application("Acme_Analyst")
        self.assertEqual(([], []), archiver.archive_due())

    def test_old_folder_is_zipped_and_removed(self):
        folder = self.make_application("Acme_Analyst")
        future = time.time() + (archiver.MAX_AGE_DAYS + 1) * 86400
        archived, _ = archiver.archive_due(now=future)
        self.assertEqual(["Acme_Analyst"], archived)
        self.assertFalse(folder.exists())
        self.assertTrue((archiver.ARCHIVE_DIR / "Acme_Analyst.zip").is_file())

    def test_dry_run_changes_nothing(self):
        folder = self.make_application("Acme_Analyst")
        future = time.time() + (archiver.MAX_AGE_DAYS + 1) * 86400
        archived, _ = archiver.archive_due(now=future, dry_run=True)
        self.assertEqual(["Acme_Analyst"], archived)
        self.assertTrue(folder.exists())

    def test_reserved_directories_are_never_archived(self):
        (self.applications / "archive").mkdir()
        (self.applications / "applied").mkdir()
        future = time.time() + (archiver.MAX_AGE_DAYS + 1) * 86400
        self.assertEqual(([], []), archiver.archive_due(now=future))

    def test_empty_submitted_date_does_not_move_a_folder(self):
        self.make_application("Rivermouth_Environmental_Data_Analyst")
        write_tracker(archiver.TRACKER_CSV, [
            {"company": "Rivermouth Environmental Consulting",
             "role": "Environmental Data Analyst", "submitted_date": ""}
        ])
        self.assertEqual([], archiver.move_applied())

    def test_filled_submitted_date_moves_the_folder(self):
        self.make_application("Rivermouth_Environmental_Data_Analyst")
        write_tracker(archiver.TRACKER_CSV, [
            {"company": "Rivermouth Environmental Consulting",
             "role": "Environmental Data Analyst",
             "submitted_date": "2026-08-01"}
        ])
        moved = archiver.move_applied()
        self.assertEqual(["Rivermouth_Environmental_Data_Analyst"], moved)
        self.assertTrue(
            (archiver.APPLIED_DIR / "Rivermouth_Environmental_Data_Analyst").is_dir()
        )

    def test_dotted_folder_name_does_not_overwrite_a_previous_zip(self):
        """REGRESSION: the one destructive bug found in the review.

        `Path("Acme_Inc._Data_Analyst").with_suffix(".zip")` yields
        `Acme_Inc.zip`, while make_archive wrote `Acme_Inc._Data_Analyst.zip`.
        The existence check and the written file were different paths, so the
        collision counter never fired and re-archiving a folder of the same name
        silently overwrote the earlier zip — after the live folder had been
        deleted. Company names ending "Inc." are the normal case.
        """
        future = time.time() + (archiver.MAX_AGE_DAYS + 1) * 86400
        name = "Acme_Inc._Data_Analyst"

        self.make_application(name)
        archiver.archive_due(now=future)
        self.make_application(name)  # a second application, same employer+role
        archiver.archive_due(now=future)

        zips = sorted(p.name for p in archiver.ARCHIVE_DIR.glob("*.zip"))
        self.assertEqual([f"{name}-2.zip", f"{name}.zip"], zips,
                         "the second archive must not overwrite the first")

    def test_an_old_file_copied_in_does_not_age_the_folder(self):
        """REGRESSION: folder age used to be the oldest mtime of its contents.

        /apply-any copies a CV and PDFs into a brand-new folder. Any copy that
        preserves mtime (shutil.copy2, a years-old CV) made the folder instantly
        "8 weeks old", so it was zipped and deleted on the very next run.
        """
        folder = self.make_application("Acme_Analyst")
        ancient = folder / "old_cv.tex"
        ancient.write_text("content", encoding="utf-8")
        long_ago = time.time() - (archiver.MAX_AGE_DAYS * 10) * 86400
        os.utime(ancient, (long_ago, long_ago))

        archived, _ = archiver.archive_due()
        self.assertEqual([], archived, "a new folder must not age via its contents")
        self.assertTrue(folder.exists())

    def test_created_marker_is_authoritative(self):
        folder = self.make_application("Acme_Analyst")
        old = (datetime.datetime.now()
               - datetime.timedelta(days=archiver.MAX_AGE_DAYS + 5))
        (folder / archiver.CREATED_MARKER).write_text(old.isoformat(),
                                                      encoding="utf-8")
        archived, _ = archiver.archive_due(dry_run=True)
        self.assertEqual(["Acme_Analyst"], archived)

    def test_open_application_is_never_archived(self):
        """An interview in progress must survive the 8-week sweep.

        Archiving deletes the live folder, and /interview prepares from exactly
        those documents.
        """
        folder = self.make_application("Rivermouth_Environmental_Data_Analyst")
        write_tracker(archiver.TRACKER_CSV, [
            {"date": "2026-01-01",
             "company": "Rivermouth Environmental Consulting",
             "role": "Environmental Data Analyst", "status": "interview_only"}
        ])
        future = time.time() + (archiver.MAX_AGE_DAYS + 1) * 86400
        archived, protected = archiver.archive_due(now=future)
        self.assertEqual([], archived)
        self.assertEqual(["Rivermouth_Environmental_Data_Analyst"], protected)
        self.assertTrue(folder.exists())

    def test_upstream_status_vocabulary_also_protects(self):
        """A row /outcome wrote as `applied` is open too."""
        self.make_application("Rivermouth_Environmental_Data_Analyst")
        write_tracker(archiver.TRACKER_CSV, [
            {"date": "2026-01-01",
             "company": "Rivermouth Environmental Consulting",
             "role": "Environmental Data Analyst", "status": "applied"}
        ])
        future = time.time() + (archiver.MAX_AGE_DAYS + 1) * 86400
        archived, protected = archiver.archive_due(now=future)
        self.assertEqual([], archived)
        self.assertEqual(["Rivermouth_Environmental_Data_Analyst"], protected)

    def test_closed_application_is_archived(self):
        """The protection must not simply disable archiving."""
        self.make_application("Rivermouth_Environmental_Data_Analyst")
        write_tracker(archiver.TRACKER_CSV, [
            {"date": "2026-01-01",
             "company": "Rivermouth Environmental Consulting",
             "role": "Environmental Data Analyst", "status": "rejected"}
        ])
        future = time.time() + (archiver.MAX_AGE_DAYS + 1) * 86400
        archived, protected = archiver.archive_due(now=future, dry_run=True)
        self.assertEqual(["Rivermouth_Environmental_Data_Analyst"], archived)
        self.assertEqual([], protected)

    def test_rows_predating_the_submitted_date_column_move_nothing(self):
        """Older CSVs have no such column; they must read as empty, not as set."""
        self.make_application("Rivermouth_Environmental_Data_Analyst")
        with archiver.TRACKER_CSV.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.writer(handle)
            writer.writerow(["date", "company", "role", "status"])
            writer.writerow(["2026-07-01", "Rivermouth Environmental Consulting",
                             "Environmental Data Analyst", "in_progress"])
        self.assertEqual([], archiver.move_applied())


if __name__ == "__main__":
    unittest.main()
